import xml.etree.cElementTree as ET
from openpyxl import load_workbook
from tools import file_operations
from config import reader_config as cfg
from testlink_integration import TestLinkUploader  # Nuevo módulo para TestLink
import os

def xl_to_xml_for_testlink(columns_in_use: dict, row_to_start: int, folder_xl_file_list: str, upload_to_testlink: bool = False):
    tl_uploader = TestLinkUploader() if upload_to_testlink else None

    for file_path in folder_xl_file_list:
        try:
            if os.path.basename(file_path).startswith('~$'):
                print(f"⚠️ Saltando archivo temporal: {file_path}")
                continue

            workbook = load_workbook(file_path)
            worksheet = workbook.active
            iter_rows = worksheet.iter_rows(min_row=row_to_start)

            main_test_suite = ET.Element('testsuite')
            current_test_suite = main_test_suite
            current_project_id = None
            test_case = None
            steps = None
            case_step = 1

            for row in iter_rows:
                cell_values = {}
                for cell in row:
                    if cell.value is not None and cell.column_letter in columns_in_use:
                        cell_values[columns_in_use[cell.column_letter]] = str(cell.value).strip()

                if 'id_proyecto' in cell_values:
                    current_project_id = cell_values['id_proyecto']
                    cfg.TESTLINK_PROJECT_ID = current_project_id

                if 'test_suite_principal' in cell_values:
                    suite_name = cell_values['test_suite_principal']
                    current_test_suite = ET.SubElement(main_test_suite, 'testsuite', name=suite_name)

                if 'caso_de_prueba' in cell_values:
                    case_name = cell_values['caso_de_prueba']
                    test_case = ET.SubElement(current_test_suite, 'testcase', name=case_name)
                    steps = ET.SubElement(test_case, 'steps')
                    case_step = 1

                if test_case is not None:
                    for field in ['importancia', 'detalle_caso', 'precondiciones']:
                        if field in cell_values:
                            xml_tag = {
                                'importancia': 'importance',
                                'detalle_caso': 'summary',
                                'precondiciones': 'preconditions'
                            }[field]
                            ET.SubElement(test_case, xml_tag).text = cell_values[field]

                    if 'pasos' in cell_values:
                        if steps is None:
                            steps = ET.SubElement(test_case, 'steps')
                        step = ET.SubElement(steps, 'step')
                        ET.SubElement(step, 'step_number').text = str(case_step)
                        ET.SubElement(step, 'actions').text = cell_values['pasos']

                    if 'resultado_esperado' in cell_values and steps is not None:
                        last_step = steps.find(f'.//step[step_number="{case_step}"]')
                        if last_step is not None:
                            ET.SubElement(last_step, 'expectedresults').text = cell_values['resultado_esperado']
                            case_step += 1
            #Guardar XML
            if len(main_test_suite) > 0:
                file_name = file_operations.prepare_file_name(file_path, "")
                xml_path = os.path.join("xml_results", f"{file_name}.xml")
                file_operations.write_tree_to_xml_file(main_test_suite, xml_path)

                if upload_to_testlink and current_project_id is not None:
                    tl_uploader.upload_from_xml(xml_path, current_project_id)

            
                # Eliminar el archivo XML después de la carga
                os.remove(xml_path)
                print(f"Archivo XML eliminado: {xml_path}")

                #mueve archivos procesados a carpeta 
                processed_dir = 'archivos_procesados'
                os.makedirs(processed_dir, exist_ok=True)
                new_path = os.path.join(processed_dir, os.path.basename(file_path))
                os.replace(file_path, new_path)

        except Exception as e:
            print(f"❌ Error procesando {file_path}: {str(e)}")


if __name__ == "__main__":
    try:
        folder_xl_file_list = file_operations.list_files(cfg.file_format)

        if not folder_xl_file_list:
            print("⚠️ No se encontraron archivos con el formato especificado")
        else:
            xl_to_xml_for_testlink(
                columns_in_use=cfg.columns_in_use,
                row_to_start=cfg.row_to_start,
                folder_xl_file_list=folder_xl_file_list,
                upload_to_testlink=True  # Cambiar a False si solo quieres el XML
            )
            print("🏁 Proceso completado exitosamente")

    except Exception as e:
        print(f"❌ Error en ejecución: {str(e)}")